"""
Given an excel file of storefront names and sellerID (columns 1 and 2, respectively),
Creates a new excel file with the SellerRatings URL for each storefront in column 3
"""

from bs4 import BeautifulSoup
import openpyxl
import requests

import datetime
import os

from utils.storefrontPageUtils import identifyKeyParagraph, getStorefrontName, getSellerID
from utils.mainPageUtils import base_urls

# function that generates possible SellerRatings url paths given a storefront name
def convertStorefrontNameToPath(storefront):
    possible_paths = []
    num_of_hyphens_at_end = 3
    new_name = [''] * (len(storefront) + num_of_hyphens_at_end)
    pointer = 0

    for idx, char in enumerate(storefront):
        # leave alone if alphabet or number
        if char.isalpha() or char.isnumeric():
            new_name[pointer] = char.lower()
            pointer += 1
        else:
            # leave alone if hyphen is the last character (unless the prior character was non-alphabet or and non-number)
            if idx == len(storefront) - 1 and char == '-' and new_name[pointer - 1] != '-':
                new_name[pointer] = '-'
                pointer += 1
            # replace consecutive spaces and/or hyphens with a single hyphen
            elif char == '-' or char == ' ':
                if new_name[pointer - 1].isalpha() or new_name[pointer - 1].isnumeric():
                    new_name[pointer] = '-'
                    pointer += 1
            # ignore consecutive non-alphabets and non-numbers

    possible_paths.append(''.join(new_name))

    for num in range(num_of_hyphens_at_end):
        new_name[pointer + num] = '-'
        possible_paths.append(''.join(new_name))

    return possible_paths


# function that finds the right url by comparing the storefront name to the one in the web page
def getRightPath(storefront, seller_ID, paths):
    valid_urls = []
    
    for base in base_urls.values():
        for path in paths:
            url = '/'.join([base, path])
            response = requests.get(url)

            if response.status_code == 200:
                soup = BeautifulSoup(response.content, "html.parser")
                key_para = identifyKeyParagraph(soup)
                
                # if sellerID matches, found exact match
                if seller_ID == getSellerID(key_para):
                    return url

                # if storefront name matches, move on to next country
                if storefront == getStorefrontName(key_para):
                    valid_urls.append(url)
                    break
    
    # if one valid url, found exact match
    if len(valid_urls) == 1:
        return valid_urls[0]
    # get user input to select one url
    elif len(valid_urls) > 1:
        user_confirm = ''

        while user_confirm == 'n' or user_confirm == '':
            user_input = ''
            user_confirm = ''
            first_try = True
            first_try_confirm = True
            
            while not user_input.isnumeric() or int(user_input) not in range(len(valid_urls)):
                if first_try:
                    print(f'For the storefront "{storefront}", there are more than one possible urls. They are as follows: ')
                    for idx, valid_url in enumerate(valid_urls):
                        print(f'{idx}: {valid_url}')
                    print('Please review and return the number corresponding to the correct url. If none of them are correct, please type in "none" or "n".')
                else:
                    print('Try again. Your input is invalid.')
                first_try = False
                user_input = input()

            while user_confirm.lower() not in {'y','n'}:
                if first_try_confirm:
                    print(f'You selected {user_input}. If it is correct, please type "y". Else, "n".')
                else:
                    print('Try again. Your input is invalid.')
                first_try_confirm = False
                user_confirm = input()
        
        if user_input.isalpha() and (user_input.lower() == "none" or user_input.lower() == "n"):
            return None
        else:
            return valid_url[user_input]

# load excel
file_path = '../data.xlsx'
workbook = openpyxl.load_workbook(file_path) # Q: should i get the user to specify the spreadsheet?
sheet = workbook.active

# create a column to add urls
sheet.insert_cols(3)
sheet['C1'] = "url"

# find each storefront's url and add to excel
for row_idx in range(2, sheet.max_row+1):
    storefront = sheet.cell(row=row_idx, column=1).value
    seller_ID = sheet.cell(row=row_idx, column=2).value
    print(f'storefront: {storefront}')
    possible_paths = convertStorefrontNameToPath(storefront)
    sheet['C'+str(row_idx)] = getRightPath(storefront, seller_ID, possible_paths)

# try to save
workbook.save(os.path.splitext(file_path)[0] + " - " + datetime.datetime.now().strftime("%m-%d-%Y %I %M %p") + '.xlsx')
